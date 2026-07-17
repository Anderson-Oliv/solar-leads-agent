import os
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook
from supabase import create_client, Client

# Instale dependências:
# pip install supabase openpyxl python-dotenv
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY")
TABLE_NAME = os.getenv("SUPABASE_TABLE", "clientes_varejistas")
BATCH_SIZE = int(os.getenv("BATCH_SIZE", "500"))

FILES = [
    "consumidores_varejistas_2W_Americana_Piracicaba_Jundiaí e Região.xlsx",
    "consumidores_varejistas_2W_São Paulo_SP.xlsx",
]

HEADER_MAP = {
    "GRP": "grp",
    "CNPJ": "cnpj",
    "TIPO__1": "tipo_empresa",
    "RAZAO_SOCIAL": "razao_social",
    "NOME_FANTASIA": "nome_fantasia",
    "SITUACAO_CADASTRAL": "situacao_cadastral",
    "DATA_SITUACAO_CADASTRAL": "data_situacao_cadastral",
    "DATA_INICIO_ATIVIDADE": "data_inicio_atividade",
    "TIPO_LOGRADOURO": "tipo_logradouro",
    "LOGRADOURO": "logradouro",
    "BAIRRO": "bairro",
    "CEP": "cep",
    "CIDADE": "cidade",
    "ESTADO": "estado",
    "NUMERO": "numero",
    "COMPLEMENTO": "complemento",
    "DDD1": "ddd1",
    "TELEFONE1": "telefone1",
    "DDD2": "ddd2",
    "TELEFONE2": "telefone2",
    "DDD_FAX": "ddd_fax",
    "FAX": "fax",
    "EMAIL": "email",
    "SITUACAO_ESPECIAL": "situacao_especial",
    "DATA_SITUACAO_ESPECIAL": "data_situacao_especial",
    "CAPITAL_SOCIAL": "capital_social",
    "RESPONSAVEL_FEDERATIVO": "responsavel_federativo",
    "CPF_CNPJ_SOCIO": "cpf_cnpj_socio",
    "NOME": "nome_socio",
    "TIPO__2": "tipo_socio",
    "CPF_REPRESENTANTE_LEGAL": "cpf_representante_legal",
    "NOME_REPRESENTANTE": "nome_representante",
    "FAIXA_ETARIA": "faixa_etaria",
    "DESCRICAO__1": "descricao_socio",
    "CNAE": "cnae",
    "DESCRICAO__2": "descricao_cnae",
}

DATE_FIELDS = {"data_situacao_cadastral", "data_inicio_atividade", "data_situacao_especial"}
NUMERIC_FIELDS = {"capital_social"}
INT_FIELDS = {"grp"}
ONLY_DIGITS_FIELDS = {"cnpj", "cep", "ddd1", "telefone1", "ddd2", "telefone2", "ddd_fax", "fax", "cpf_cnpj_socio", "cpf_representante_legal", "cnae"}


def normalize_header(raw_headers: Iterable[Any]) -> List[str]:
    """Resolve colunas repetidas como TIPO e DESCRICAO."""
    counts: Dict[str, int] = {}
    result = []
    for h in raw_headers:
        base = str(h or "").strip().upper()
        counts[base] = counts.get(base, 0) + 1
        if counts[base] == 1 and base not in {"TIPO", "DESCRICAO"}:
            result.append(base)
        else:
            result.append(f"{base}__{counts[base]}")
    return result


def only_digits(value: Any) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits or None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    txt = str(value).strip()
    return txt if txt else None


def excel_date_to_iso(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date: day 1 = 1899-12-31, with Excel's 1900 leap-year bug.
        from openpyxl.utils.datetime import from_excel
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def to_decimal(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return str(Decimal(text))
    except InvalidOperation:
        return None


def build_derived_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    ddd = row.get("ddd1")
    tel = row.get("telefone1")
    row["telefone_principal"] = f"{ddd}{tel}" if ddd and tel else None

    partes = [
        row.get("tipo_logradouro"), row.get("logradouro"), row.get("numero"),
        row.get("complemento"), row.get("bairro"), row.get("cidade"), row.get("estado"), row.get("cep")
    ]
    row["endereco_completo"] = ", ".join([str(p) for p in partes if p]) or None
    return row


def normalize_record(raw: Dict[str, Any], fonte_arquivo: str) -> Dict[str, Any] | None:
    record: Dict[str, Any] = {"fonte_arquivo": fonte_arquivo}

    for source_col, target_col in HEADER_MAP.items():
        value = raw.get(source_col)
        if target_col in DATE_FIELDS:
            record[target_col] = excel_date_to_iso(value)
        elif target_col in NUMERIC_FIELDS:
            record[target_col] = to_decimal(value)
        elif target_col in INT_FIELDS:
            record[target_col] = int(value) if value not in (None, "") else None
        elif target_col in ONLY_DIGITS_FIELDS:
            record[target_col] = only_digits(value)
        else:
            record[target_col] = clean_text(value)

    if not record.get("cnpj"):
        return None

    return build_derived_fields(record)


def iter_excel_records(file_path: Path) -> Iterable[Dict[str, Any]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = normalize_header(next(rows))

    for row in rows:
        if not row or all(v is None for v in row):
            continue
        raw = dict(zip(headers, row))
        normalized = normalize_record(raw, file_path.name)
        if normalized:
            yield normalized


def chunks(items: Iterable[Dict[str, Any]], size: int) -> Iterable[List[Dict[str, Any]]]:
    batch = []
    for item in items:
        batch.append(item)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def import_files(base_dir: str = ".") -> None:
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Configure SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY no arquivo .env")

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    total = 0

    for filename in FILES:
        file_path = Path(base_dir) / filename
        if not file_path.exists():
            print(f"Arquivo não encontrado, pulando: {file_path}")
            continue

        print(f"Importando: {file_path.name}")
        for batch in chunks(iter_excel_records(file_path), BATCH_SIZE):
            # on_conflict usa a constraint unique(cnpj) criada no SQL.
            supabase.table(TABLE_NAME).upsert(batch, on_conflict="cnpj").execute()
            total += len(batch)
            print(f"  {total} registros enviados...")

    print(f"Importação finalizada. Registros enviados para upsert: {total}")


if __name__ == "__main__":
    # Execute na mesma pasta onde estão as planilhas.
    import_files(base_dir=".")
